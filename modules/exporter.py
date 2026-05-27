"""
exporter.py — Excel export with exact spec compliance.

Key rules from spec:
- Sheet name = <ClinicName_District_ready>
- No autofilter
- Vertical alignment: center
- White background on all cells
- Dynamic column widths based on content
- Arial 10, thin black borders, wrap text
- No merged cells, no colored headers

Price rule:
- If price == "Цена по запросу" (or variants) → store as "9999999" in all exports
- If any service has 9999999 price → clinic name gets * suffix (e.g. "Green Lukas*")
- * suffix means unsigned clinic with no real pricing
- Rows with ID == "-" are excluded from ready file

Description lookup priority:
- 1) descriptions_catalog.json by service ID (pre-generated)
- 2) Template fallback (modules/templates.py)

Name translation:
- Имя UZ = OpenRouter/DeepSeek translation of Имя RU
- Имя KR = transliteration of Имя UZ to Uzbek Cyrillic
"""
import io
import json
import os
import re
import time
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from modules.utils import clean_cell, safe_str, uz_latin_to_cyrillic
from modules.extractor import _normalize_district
from modules.templates import get_descriptions, get_duration
from modules.validation import REQUIRED_COLUMNS

CATEGORY_MAP = {
    "Анализы": {
        "Категория RU": "Анализы",
        "Категория UZ": "Tahlillar",
        "Категория KR": "Таҳлиллар",
    },
    "Диагностика": {
        "Категория RU": "Диагностика",
        "Категория UZ": "Diagnostika",
        "Категория KR": "Диагностика",
    },
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

CENTER_COLS = {
    "Услуга ID", "Цена", "Продолжительность (минут)",
    "Активен", "Тип услуг",
    "Код уз", "Код ру", "Код лаборатория",
}

COL_MIN_WIDTH = 8
COL_MAX_WIDTH = 60
WIDE_COLS = {
    "Имя RU", "Имя UZ", "Имя KR",
    "Описание RU", "Описание UZ", "Описание KR",
    "Требования RU", "Требования UZ", "Требования KR",
}
WIDE_COLS_MIN = 35
WIDE_COLS_MAX = 55

PRICE_ON_REQUEST_SENTINEL = "9999999"
PRICE_ON_REQUEST_LABELS   = {"цена по запросу", "по запросу", "price on request"}

OPENROUTER_URL   = "https://openrouter.ai/api/v1/chat/completions"
OPENROUTER_MODEL = "deepseek/deepseek-chat"


# ── Name translation via OpenRouter ──────────────────────────────────────────

def _translate_names_batch(names: list[str]) -> dict:
    """
    Translate a list of Russian medical service names to Uzbek Latin.
    Returns {name_ru: name_uz} dict.
    Falls back to original name on failure.
    """
    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key:
        print("[translate] No OPENROUTER_API_KEY — skipping translation")
        return {}

    results = {}
    batch_size = 100

    system_prompt = """Ты медицинский переводчик. Переводи названия медицинских услуг с русского на узбекский латинский алфавит.
Отвечай ТОЛЬКО валидным JSON массивом. Никакого текста до или после. Формат:
[{"ru":"...", "uz":"..."}]
Правила:
- Переводи точно, сохраняй медицинские термины
- Латинские аббревиатуры (МРТ, КТ, УЗИ, ПЦР, ИФА) оставляй как есть или используй общепринятый узбекский вариант
- Не добавляй лишних слов"""

    for i in range(0, len(names), batch_size):
        batch = names[i:i + batch_size]
        items = "\n".join(f'{{"ru":"{n}"}}' for n in batch)
        user_prompt = f"Переведи эти названия на узбекский латинский:\n{items}"

        for attempt in range(3):
            try:
                resp = requests.post(
                    OPENROUTER_URL,
                    headers={
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json",
                        "HTTP-Referer": "https://medpay-automation.streamlit.app",
                        "X-Title": "MedPay Name Translator",
                    },
                    json={
                        "model": OPENROUTER_MODEL,
                        "messages": [
                            {"role": "system", "content": system_prompt},
                            {"role": "user",   "content": user_prompt},
                        ],
                        "temperature": 0.1,
                        "max_tokens": 6000,
                    },
                    timeout=30,
                )
                if resp.status_code == 429:
                    time.sleep(30 * (attempt + 1))
                    continue
                if resp.status_code == 402:
                    print("[translate] OpenRouter credits depleted")
                    return results
                resp.raise_for_status()
                content = resp.json()["choices"][0]["message"]["content"].strip()
                if content.startswith("```"):
                    content = content.split("```")[1]
                    if content.startswith("json"):
                        content = content[4:]
                    content = content.strip()
                parsed = json.loads(content)
                for entry in parsed:
                    ru = entry.get("ru", "").strip()
                    uz = entry.get("uz", "").strip()
                    if ru and uz:
                        results[ru] = uz
                break
            except Exception as e:
                print(f"[translate] Attempt {attempt+1} failed: {e}")
                if attempt < 2:
                    time.sleep(5 * (attempt + 1))

        time.sleep(0.5)

    print(f"[translate] Translated {len(results)}/{len(names)} names")
    return results


# ── Descriptions catalog loader ───────────────────────────────────────────────

def _load_desc_catalog() -> dict:
    """Load pre-generated descriptions catalog from HuggingFace or local fallback."""
    local_paths = [
        "/content/drive/MyDrive/Medpay Automation/descriptions_catalog.json",
        "descriptions_catalog.json",
        "data/descriptions_catalog.json",
    ]
    for path in local_paths:
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                print(f"[desc_catalog] Loaded {len(data)} entries from {path}")
                return data
            except Exception as e:
                print(f"[desc_catalog] Failed to load {path}: {e}")

    hf_token = os.environ.get("HF_TOKEN", "")
    hf_url = "https://huggingface.co/admin11011/medpay-matcher/resolve/main/descriptions_catalog.json"
    try:
        print("[desc_catalog] Downloading from HuggingFace...")
        headers = {"Authorization": f"Bearer {hf_token}"} if hf_token else {}
        resp = requests.get(hf_url, headers=headers, timeout=120)
        resp.raise_for_status()
        data = resp.json()
        print(f"[desc_catalog] Downloaded {len(data)} entries from HuggingFace")
        try:
            with open("descriptions_catalog.json", "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False)
            print("[desc_catalog] Cached locally")
        except Exception:
            pass
        return data
    except Exception as e:
        print(f"[desc_catalog] HuggingFace download failed: {e}")

    print("[desc_catalog] All sources failed — using template fallback")
    return {}


# ── Price helpers ─────────────────────────────────────────────────────────────

def _normalize_price(price) -> str:
    val = str(price).strip()
    if val.lower() in PRICE_ON_REQUEST_LABELS or val == PRICE_ON_REQUEST_SENTINEL:
        return PRICE_ON_REQUEST_SENTINEL
    return val


def _has_on_request_price(rows: list[dict]) -> bool:
    for row in rows:
        if _normalize_price(row.get("Цена", "")) == PRICE_ON_REQUEST_SENTINEL:
            return True
    return False


def _clinic_export_name(clinic_name: str, has_on_request: bool) -> str:
    name = clean_cell(clinic_name.strip().rstrip("*").strip())
    if has_on_request:
        return name + "*"
    return name


def _make_sheet_name(clinic_name: str, district: str) -> str:
    clean_cn = clinic_name.rstrip("*").strip()
    base = f"{clean_cell(clean_cn)}_{clean_cell(district)}"
    safe = re.sub(r"[\\/*?\[\]:]", "_", base)
    return safe[:31] if safe else "Ready"


def _calc_col_width(ws, col_idx: int, col_name: str, max_rows: int = 100) -> float:
    max_len = len(col_name)
    for row in range(2, min(ws.max_row + 1, max_rows + 2)):
        val = ws.cell(row=row, column=col_idx).value
        if val is not None:
            text = str(val)
            effective = min(len(text), max(
                max(len(word) for word in text.split()) + 2,
                len(text) // 3
            ))
            max_len = max(max_len, effective)
    if col_name in WIDE_COLS:
        return max(WIDE_COLS_MIN, min(WIDE_COLS_MAX, max_len + 4))
    return max(COL_MIN_WIDTH, min(COL_MAX_WIDTH, max_len + 4))


def _apply_cell_style(cell, is_header: bool = False, col_name: str = ""):
    cell.font = Font(name="Arial", size=10, bold=is_header)
    cell.fill = WHITE_FILL
    cell.border = THIN_BORDER
    if is_header:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        h_align = "center" if col_name in CENTER_COLS else "left"
        cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)


# ── Main builder ──────────────────────────────────────────────────────────────

def build_ready_df(
    matched_rows: list[dict],
    clinic_name: str,
    district: str,
    catalog_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build final ready DataFrame with all 22 columns."""

    # Load descriptions catalog
    try:
        desc_catalog = _load_desc_catalog()
    except Exception as e:
        print(f"[desc_catalog] Load error: {e}")
        desc_catalog = {}
    print(f"[desc_catalog] Entries loaded: {len(desc_catalog)}")

    # Normalize prices
    for row in matched_rows:
        row["Цена"] = _normalize_price(row.get("Цена", ""))

    has_on_request = _has_on_request_price(matched_rows)
    export_name    = _clinic_export_name(clinic_name, has_on_request)
    district       = _normalize_district(district.strip())

    # Collect all unique clinic service names for translation
    unique_names = list({
        clean_cell(str(row.get("Название в клинике", "")))
        for row in matched_rows
        if clean_cell(str(row.get("ID", "-"))) not in ("-", "")
        and clean_cell(str(row.get("Название в клинике", "")))
    })
    print(f"[translate] Translating {len(unique_names)} unique service names...")
    name_translations = _translate_names_batch(unique_names)

    records = []

    for row in matched_rows:
        service_id = clean_cell(str(row.get("ID", "-")))
        if service_id in ("-", ""):
            continue

        service_type = clean_cell(str(row.get("Тип услуг", "Диагностика")))
        clinic_svc   = clean_cell(str(row.get("Название в клинике", "")))
        price        = row.get("Цена", PRICE_ON_REQUEST_SENTINEL)

        if service_type not in ("Диагностика", "Анализы"):
            service_type = "Диагностика"

        # Catalog entry for descriptions
        catalog_entry = desc_catalog.get(str(service_id), {})

        # ── Names ─────────────────────────────────────────────────────────────
        name_ru = clinic_svc  # always use clinic's original RU name

        # Имя UZ = OpenRouter translation of clinic's RU name
        name_uz = name_translations.get(clinic_svc, "")
        if not name_uz:
            # Fallback: catalog UZ name
            cat_row = catalog_df[catalog_df["ID number"] == service_id]
            if not cat_row.empty:
                cat_uz = safe_str(cat_row.iloc[0].get("Name UZ", ""))
                if cat_uz and len(cat_uz) > 2:
                    name_uz = cat_uz
            if not name_uz:
                name_uz = catalog_entry.get("name_uz", "") or clinic_svc

        # Имя KR = transliteration of Имя UZ
        name_kr = uz_latin_to_cyrillic(name_uz) if name_uz else clinic_svc

        name_ru = clean_cell(name_ru)
        name_uz = clean_cell(name_uz)
        name_kr = clean_cell(name_kr)

        # ── Descriptions and requirements ─────────────────────────────────────
        if catalog_entry and catalog_entry.get("desc_ru", "").strip():
            desc_ru = clean_cell(catalog_entry.get("desc_ru", ""))
            desc_uz = clean_cell(catalog_entry.get("desc_uz", ""))
            desc_kr = clean_cell(catalog_entry.get("desc_kr", ""))
            req_ru  = clean_cell(catalog_entry.get("req_ru", ""))
            req_uz  = clean_cell(catalog_entry.get("req_uz", ""))
            req_kr  = clean_cell(uz_latin_to_cyrillic(req_uz)) if req_uz else ""
            if not desc_kr and desc_uz:
                desc_kr = clean_cell(uz_latin_to_cyrillic(desc_uz))
            texts = {
                "Описание RU":   desc_ru,
                "Описание UZ":   desc_uz,
                "Описание KR":   desc_kr,
                "Требования RU": req_ru,
                "Требования UZ": req_uz,
                "Требования KR": req_kr,
            }
        else:
            texts = get_descriptions(service_type, name_ru, name_uz, name_kr)

        duration = get_duration(service_type, name_ru)
        cat      = CATEGORY_MAP.get(service_type, CATEGORY_MAP["Диагностика"])

        records.append({
            "Услуга ID":                  service_id,
            "Клиника":                    export_name,
            "Филиал":                     district,
            "Имя RU":                     name_ru,
            "Имя UZ":                     name_uz,
            "Имя KR":                     name_kr,
            "Описание UZ":                clean_cell(texts.get("Описание UZ", "")),
            "Требования UZ":              clean_cell(texts.get("Требования UZ", "")),
            "Требования RU":              clean_cell(texts.get("Требования RU", "")),
            "Требования KR":              clean_cell(texts.get("Требования KR", "")),
            "Тип услуг":                  service_type,
            "Цена":                       price,
            "Описание RU":                clean_cell(texts.get("Описание RU", "")),
            "Описание KR":                clean_cell(texts.get("Описание KR", "")),
            "Продолжительность (минут)":  str(duration),
            "Активен":                    "TRUE",
            "Категория RU":               cat["Категория RU"],
            "Категория UZ":               cat["Категория UZ"],
            "Категория KR":               cat["Категория KR"],
            "Код уз":                     "",
            "Код ру":                     "",
            "Код лаборатория":            "",
        })

    return pd.DataFrame(records, columns=REQUIRED_COLUMNS)


# ── Excel exporters ───────────────────────────────────────────────────────────

def export_price_list_excel(
    services: list[dict], clinic_name: str, district: str
) -> bytes:
    rows = [
        {
            "#": i,
            "Название услуги": s.get("service_name", ""),
            "Тип": s.get("type", ""),
            "Цена": _normalize_price(s.get("price", "")),
        }
        for i, s in enumerate(services, 1)
    ]
    df  = pd.DataFrame(rows)
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Прайс-лист"

    for ci, col in enumerate(df.columns, 1):
        _apply_cell_style(ws.cell(row=1, column=ci, value=col), is_header=True)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci,
                           value=None if str(val) in ("", "nan") else val)
            _apply_cell_style(cell, col_name=df.columns[ci - 1])
    for ci, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _calc_col_width(ws, ci, col)
    ws.freeze_panes = "A2"
    wb.save(buf)
    return buf.getvalue()


def export_matching_excel(matched_rows: list[dict]) -> bytes:
    cols = ["Название в MedPay", "Название в клинике", "ID",
            "Уверенность", "Комментарий", "Цена", "Тип услуг"]
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Матчинг"

    red_fill    = PatternFill("solid", fgColor="FFE0E0")
    yellow_fill = PatternFill("solid", fgColor="FFFDE0")

    for ci, col in enumerate(cols, 1):
        _apply_cell_style(ws.cell(row=1, column=ci, value=col), is_header=True)
    for ri, rec in enumerate(matched_rows, 2):
        id_val       = str(rec.get("ID", "")).strip()
        conf         = rec.get("Уверенность", 0)
        is_unmatched = id_val == "-"
        is_low_conf  = not is_unmatched and isinstance(conf, (int, float)) and conf < 90
        for ci, col in enumerate(cols, 1):
            val = rec.get(col, "")
            if col == "Цена":
                val = _normalize_price(val)
            cell = ws.cell(row=ri, column=ci,
                           value=None if str(val) in ("", "nan") else val)
            _apply_cell_style(cell, col_name=col)
            if is_unmatched:
                cell.fill = red_fill
            elif is_low_conf:
                cell.fill = yellow_fill
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _calc_col_width(ws, ci, col)
    ws.freeze_panes = "A2"
    wb.save(buf)
    return buf.getvalue()


def export_ready_excel(
    ready_df: pd.DataFrame,
    clinic_name: str = "",
    district: str = "",
) -> bytes:
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active

    clean_cn   = clinic_name.strip().rstrip("*").strip()
    sheet_name = _make_sheet_name(clean_cn, district) + "_ready"
    ws.title   = sheet_name[:31]

    for ci, col in enumerate(REQUIRED_COLUMNS, 1):
        _apply_cell_style(ws.cell(row=1, column=ci, value=col), is_header=True, col_name=col)
    for ri in range(len(ready_df)):
        rec = ready_df.iloc[ri]
        for ci, col in enumerate(REQUIRED_COLUMNS, 1):
            raw = rec[col]
            if raw is None or str(raw).strip().lower() in ("nan", "none", ""):
                val = None
            elif col in ("Код уз", "Код ру", "Код лаборатория"):
                val = None
            elif col == "Цена":
                val = _normalize_price(str(raw))
            else:
                val = clean_cell(str(raw)) or None
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            _apply_cell_style(cell, col_name=col)
    for ci, col in enumerate(REQUIRED_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _calc_col_width(ws, ci, col)
    ws.row_dimensions[1].height = 25
    for ri in range(2, ws.max_row + 1):
        max_chars = max(
            (len(str(ws.cell(row=ri, column=ci).value or ""))
             for ci in range(1, 23)),
            default=0,
        )
        estimated_lines = max(1, max_chars // 40)
        ws.row_dimensions[ri].height = max(20, min(120, estimated_lines * 15))
    ws.freeze_panes = "A2"
    wb.save(buf)
    return buf.getvalue()
